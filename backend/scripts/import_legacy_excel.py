"""
One-time import of the legacy "Monthly Expenses" Excel workbook.

The workbook holds one sheet per fiscal year (Apr..Mar) with monthly totals
per line item, an Income row, and per-month Remarks. This script converts
each (line item, month) cell into a single transaction dated the 1st of the
month, each Income cell into an income entry, and each remark into a monthly
remark — so the dashboard matrix reproduces the old spreadsheet exactly.

    python -m scripts.import_legacy_excel --file "path/to/Monthly Expenses.xlsx" \
        --username sreenath [--dry-run]

Runs against whatever DATABASE_URL the app is configured with, so the same
command works on local SQLite and (with DATABASE_URL pointed at Neon) on
production Postgres.

Behavior notes:
  - Blocks/line items are created as the sheets label them (bold rows are
    block headers). Line items that don't already exist for the user are
    created ARCHIVED so historical categories don't clutter the dropdowns.
  - FY 22-23 has no block labels; items are matched by name against the
    labeled sheets, falling back to the "Personal" block.
  - Valued rows with no label are imported as an "Adjustment" line item so
    monthly totals still reconcile with the sheet's Sub-Total row.
  - The "Actual Invest Block" tables are ignored (the Summary rows reconcile
    against the first Invest Block, so that's the one imported).
  - Months where the user already has app data are skipped unless listed in
    --force-months. Previously imported rows (recognized by the description
    marker) are replaced on re-run, so the script is idempotent.
  - Every transaction gets payment source --payment-source (default
    "Debit Card"); every income entry gets income source --income-source
    (default "Debit Card", per the import request).
  - After writing, per-month totals are reconciled against the sheets'
    Sub-Total rows and mismatches are reported.
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from datetime import date
from decimal import Decimal

import openpyxl
from sqlalchemy import extract, func, or_, select

from app.database import SessionLocal
from app.models import (
    Block,
    IncomeEntry,
    IncomeSource,
    LineItem,
    MonthlyRemark,
    PaymentSource,
    Transaction,
    User,
)

MARKER = "Imported from Excel"
MONTHS = {
    "apr": 4, "may": 5, "jun": 6, "jul": 7, "aug": 8, "sep": 9,
    "oct": 10, "nov": 11, "dec": 12, "jan": 1, "feb": 2, "mar": 3,
}
SECTION_LABELS = {
    "expense block", "invest block", "summary", "total expenditure summary",
    "total exp summary", "actual invest block",
}
FALLBACK_BLOCK = "Personal"
ADJUSTMENT_ITEM = "Adjustment"


def norm(s: str) -> str:
    return re.sub(r"\s+", " ", s).strip().lower()


def is_bold(cell) -> bool:
    return bool(cell.value is not None and cell.font and cell.font.bold)


def fiscal_start_year(sheet_name: str) -> int:
    m = re.search(r"(\d{2})", sheet_name)
    if not m:
        raise ValueError(f"Cannot parse fiscal year from sheet name {sheet_name!r}")
    return 2000 + int(m.group(1))


def cell_amount(value) -> Decimal | None:
    """Numeric cell → Decimal; blanks/text/zero → None (nothing to import)."""
    if value is None or isinstance(value, str):
        return None
    d = Decimal(str(round(float(value), 2)))
    return d if d != 0 else None


def parse_sheet(ws):
    """Extract the sheet's structure. Returns a dict with:
    months: {col_idx: (year, month)}, expense/invest item rows, income,
    remarks, and the sheet's own Sub-Total rows for reconciliation."""
    start_year = fiscal_start_year(ws.title)

    # Month columns from row 1 (Apr..Mar; Jan-Mar roll into the next year).
    months: dict[int, tuple[int, int]] = {}
    for cell in ws[1]:
        if isinstance(cell.value, str) and norm(cell.value) in MONTHS:
            m = MONTHS[norm(cell.value)]
            months[cell.column] = (start_year if m >= 4 else start_year + 1, m)
    if not months:
        raise ValueError(f"{ws.title}: no month headers found in row 1")

    # Layout: the 'Expense Block' label sits in the item-name column in the
    # old layout (blocks share that column, bold) and one column left of the
    # items in the new layout (blocks in B, items in C).
    anchor = None
    for row in ws.iter_rows(max_col=4):
        for c in row:
            if isinstance(c.value, str) and norm(c.value) == "expense block":
                anchor = c
                break
        if anchor:
            break
    if anchor is None:
        raise ValueError(f"{ws.title}: 'Expense Block' label not found")
    # New layout (FY 25-26 onward): block names sit bold in the column LEFT of
    # the 'Expense Block' label, items in the label's column. Old layout:
    # blocks and items share the label's column (bold rows are blocks).
    new_layout = anchor.column > 1 and any(
        is_bold(row[0]) and isinstance(row[0].value, str)
        for row in ws.iter_rows(
            min_col=anchor.column - 1, max_col=anchor.column - 1
        )
    )
    item_col = anchor.column
    block_col = anchor.column - 1 if new_layout else anchor.column

    def label(row_idx: int, col: int):
        v = ws.cell(row=row_idx, column=col).value
        return v.strip() if isinstance(v, str) and v.strip() else None

    def row_values(row_idx: int) -> dict[tuple[int, int], Decimal]:
        out = {}
        for col, ym in months.items():
            amt = cell_amount(ws.cell(row=row_idx, column=col).value)
            if amt is not None:
                out[ym] = amt
        return out

    expense_items: list[tuple[str | None, str, dict]] = []  # (block, item, values)
    invest_items: list[tuple[str | None, str, dict]] = []
    subtotals: dict[str, dict[tuple[int, int], Decimal]] = {}
    income: dict[tuple[int, int], Decimal] = {}
    remarks: dict[tuple[int, int], str] = {}
    unlabeled: list[int] = []

    section = None  # None | 'expense' | 'invest' | 'done'
    current_block: str | None = None
    for row_idx in range(1, ws.max_row + 1):
        item_label = label(row_idx, item_col)
        block_label = label(row_idx, block_col) if new_layout else None
        item_cell = ws.cell(row=row_idx, column=item_col)
        block_cell = ws.cell(row=row_idx, column=block_col)

        # Section transitions keyed off the labels present on this row.
        labels_here = {norm(x) for x in (item_label, block_label) if x}
        if "expense block" in labels_here:
            section = "expense"
            # In the new layout the first item shares the next row already.
            continue
        if "invest block" in labels_here and section != "done":
            section = "invest"
            current_block = "Invest Block"
            # New layout: first invest item may share this row.
            if new_layout and item_label and norm(item_label) != "invest block":
                vals = row_values(row_idx)
                if vals:
                    invest_items.append(("Invest Block", item_label, vals))
            continue
        if labels_here & {"summary", "total expenditure summary", "total exp summary"}:
            section = "done"
        if labels_here & {"actual invest block"}:
            section = "done"

        # Income / Remarks rows (they appear after the sections).
        if item_label and norm(item_label) == "income" and not income:
            income = row_values(row_idx)
            continue
        if item_label and norm(item_label) == "remarks" and not remarks:
            for col, ym in months.items():
                v = ws.cell(row=row_idx, column=col).value
                if isinstance(v, str) and v.strip():
                    remarks[ym] = v.strip()
            continue

        if section not in ("expense", "invest"):
            continue

        # Sub-Total ends the current section.
        if item_label and norm(item_label) == "sub-total":
            subtotals[section] = row_values(row_idx)
            section = None if section == "expense" else "done"
            continue

        # Block headers: bold cell in the block column (new layout) or a bold
        # non-item row in the shared column (old layout).
        if new_layout and block_label and is_bold(block_cell):
            current_block = block_label
        if not new_layout and item_label and is_bold(item_cell):
            if norm(item_label) not in SECTION_LABELS:
                current_block = item_label
            continue

        vals = row_values(row_idx)
        if item_label:
            target = expense_items if section == "expense" else invest_items
            block = "Invest Block" if section == "invest" else current_block
            target.append((block, item_label, vals))
        elif vals:
            # Valued row with no label — keep totals honest via 'Adjustment'.
            unlabeled.append(row_idx)
            target = expense_items if section == "expense" else invest_items
            block = "Invest Block" if section == "invest" else current_block
            target.append((block, ADJUSTMENT_ITEM, vals))

    return {
        "sheet": ws.title,
        "expense": expense_items,
        "invest": invest_items,
        "income": income,
        "remarks": remarks,
        "subtotals": subtotals,
        "unlabeled": unlabeled,
    }


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--file", required=True, help="Path to the legacy .xlsx workbook")
    ap.add_argument("--username", required=True, help="App account to import into")
    ap.add_argument("--payment-source", default="Debit Card")
    ap.add_argument("--income-source", default="Debit Card")
    ap.add_argument(
        "--force-months",
        default="",
        help="Comma-separated YYYY-MM months to import even if the account "
        "already has data in them",
    )
    ap.add_argument("--dry-run", action="store_true", help="Report only; no writes")
    args = ap.parse_args()

    forced = {m.strip() for m in args.force_months.split(",") if m.strip()}
    wb = openpyxl.load_workbook(args.file, data_only=True)
    parsed = [parse_sheet(ws) for ws in wb.worksheets]

    # Item-name → block map from the labeled sheets, for sheets without labels.
    name_to_block: dict[str, str] = {}
    for p in parsed:
        for block, item, _ in p["expense"]:
            if block and norm(item) not in name_to_block:
                name_to_block[norm(item)] = block

    db = SessionLocal()
    try:
        user = db.scalar(select(User).where(User.username == args.username))
        if user is None:
            sys.exit(f"No user named {args.username!r} in this database.")

        # Months that already contain non-imported app data are skipped.
        existing_months = {
            f"{int(y):04d}-{int(m):02d}"
            for y, m in db.execute(
                select(
                    extract("year", Transaction.txn_date),
                    extract("month", Transaction.txn_date),
                )
                .where(
                    Transaction.user_id == user.id,
                    # NULL descriptions are manual entries too — NOT LIKE
                    # alone would drop them (NULL comparison semantics).
                    or_(
                        Transaction.description.is_(None),
                        ~Transaction.description.like(f"{MARKER}%"),
                    ),
                )
                .distinct()
            )
        }
        skipped_months = sorted(existing_months - forced)

        def month_ok(ym: tuple[int, int]) -> bool:
            key = f"{ym[0]:04d}-{ym[1]:02d}"
            return key in forced or key not in existing_months

        # Wipe previously imported rows so re-runs replace instead of duplicate.
        prior_txns = db.scalars(
            select(Transaction).where(
                Transaction.user_id == user.id,
                Transaction.description.like(f"{MARKER}%"),
            )
        ).all()
        prior_income = db.scalars(
            select(IncomeEntry).where(
                IncomeEntry.user_id == user.id,
                IncomeEntry.description.like(f"{MARKER}%"),
            )
        ).all()
        for row in [*prior_txns, *prior_income]:
            db.delete(row)

        def get_or_create_block(name: str, type_: str) -> Block:
            b = db.scalar(
                select(Block).where(Block.user_id == user.id, Block.name == name)
            )
            if b is None:
                max_sort = db.scalar(
                    select(func.max(Block.sort_order)).where(Block.user_id == user.id)
                )
                b = Block(
                    user_id=user.id,
                    name=name,
                    type=type_,
                    sort_order=(max_sort or 0) + 1,
                )
                db.add(b)
                db.flush()
                created_blocks.append(name)
            return b

        def get_or_create_item(block: Block, name: str) -> LineItem:
            for li in db.scalars(
                select(LineItem).where(LineItem.block_id == block.id)
            ):
                if norm(li.name) == norm(name):
                    return li
            max_sort = db.scalar(
                select(func.max(LineItem.sort_order)).where(
                    LineItem.block_id == block.id
                )
            )
            li = LineItem(
                user_id=user.id,
                block_id=block.id,
                name=name,
                sort_order=(max_sort or 0) + 1,
                archived=True,  # legacy-only categories stay out of dropdowns
            )
            db.add(li)
            db.flush()
            created_items.append(f"{block.name} / {name}")
            return li

        pay = db.scalar(
            select(PaymentSource).where(
                PaymentSource.user_id == user.id,
                PaymentSource.name == args.payment_source,
            )
        )
        if pay is None:
            pay = PaymentSource(user_id=user.id, name=args.payment_source)
            db.add(pay)
            db.flush()
        inc_src = db.scalar(
            select(IncomeSource).where(
                IncomeSource.user_id == user.id,
                IncomeSource.name == args.income_source,
            )
        )
        if inc_src is None:
            inc_src = IncomeSource(user_id=user.id, name=args.income_source)
            db.add(inc_src)
            db.flush()

        created_blocks: list[str] = []
        created_items: list[str] = []
        totals = {"txns": 0, "income": 0, "remarks": 0}
        imported_totals: dict[str, dict[tuple[int, int], Decimal]] = {
            "expense": defaultdict(Decimal),
            "invest": defaultdict(Decimal),
        }

        for p in parsed:
            marker = f"{MARKER} ({p['sheet']})"
            for section in ("expense", "invest"):
                for block_name, item_name, values in p[section]:
                    values = {ym: v for ym, v in values.items() if month_ok(ym)}
                    if not values:
                        continue
                    if block_name is None:
                        block_name = name_to_block.get(norm(item_name), FALLBACK_BLOCK)
                    block = get_or_create_block(
                        block_name,
                        "INVESTMENT" if section == "invest" else "EXPENSE",
                    )
                    item = get_or_create_item(block, item_name)
                    for (y, m), amount in values.items():
                        db.add(
                            Transaction(
                                user_id=user.id,
                                txn_date=date(y, m, 1),
                                amount=amount,
                                line_item_id=item.id,
                                payment_source_id=pay.id,
                                description=f"{marker} — {item_name} monthly total",
                            )
                        )
                        totals["txns"] += 1
                        imported_totals[section][(y, m)] += amount

            for (y, m), amount in p["income"].items():
                if not month_ok((y, m)):
                    continue
                db.add(
                    IncomeEntry(
                        user_id=user.id,
                        entry_date=date(y, m, 1),
                        amount=amount,
                        income_source_id=inc_src.id,
                        description=f"{marker} — monthly income total",
                    )
                )
                totals["income"] += 1

            for (y, m), body in p["remarks"].items():
                if not month_ok((y, m)):
                    continue
                exists = db.scalar(
                    select(MonthlyRemark).where(
                        MonthlyRemark.user_id == user.id,
                        MonthlyRemark.year == y,
                        MonthlyRemark.month == m,
                    )
                )
                if exists is None:
                    db.add(MonthlyRemark(user_id=user.id, year=y, month=m, body=body))
                    totals["remarks"] += 1

        # --- Reconcile against the sheets' own Sub-Total rows ---------------
        mismatches = []
        for p in parsed:
            for section in ("expense", "invest"):
                for ym, sheet_total in p["subtotals"].get(section, {}).items():
                    if not month_ok(ym):
                        continue
                    ours = imported_totals[section][ym]
                    if abs(ours - sheet_total) > Decimal("0.01"):
                        mismatches.append(
                            f"  {p['sheet']} {section} {ym[0]}-{ym[1]:02d}: "
                            f"sheet={sheet_total} imported={ours}"
                        )

        print(f"\n=== Import summary ({'DRY RUN' if args.dry_run else 'written'}) ===")
        print(f"user: {user.username}   file: {args.file}")
        if prior_txns or prior_income:
            print(
                f"replaced prior import: {len(prior_txns)} txns, "
                f"{len(prior_income)} income entries"
            )
        print(
            f"created: {totals['txns']} transactions, {totals['income']} income "
            f"entries, {totals['remarks']} remarks"
        )
        if created_blocks:
            print(f"new blocks: {', '.join(created_blocks)}")
        if created_items:
            print(f"new line items (created archived): {len(created_items)}")
            for name in created_items:
                print(f"  - {name}")
        if skipped_months:
            print(
                f"skipped months with existing app data: {', '.join(skipped_months)}"
                "  (use --force-months to include)"
            )
        for p in parsed:
            if p["unlabeled"]:
                print(
                    f"note: {p['sheet']} rows {p['unlabeled']} had values but no "
                    f"label — imported as '{ADJUSTMENT_ITEM}'"
                )
        if mismatches:
            print("RECONCILIATION MISMATCHES (imported vs sheet Sub-Total):")
            print("\n".join(mismatches))
        else:
            print("reconciliation: all monthly totals match the sheet Sub-Totals [OK]")

        if args.dry_run:
            db.rollback()
        else:
            db.commit()
    finally:
        db.close()


if __name__ == "__main__":
    main()
