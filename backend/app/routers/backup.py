"""Export / import a user's entire dataset as a multi-sheet .xlsx workbook.

Export produces a human-readable Excel file (openable in Excel/Sheets) that
is ALSO a complete backup. Import reads that same file back and restores the
data into the current account, finding-or-creating categories/sources by name
so a fresh account is rebuilt faithfully. Import is additive — intended for
restoring into an empty (or newly created) account.
"""

from datetime import date, datetime
from io import BytesIO

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import (
    Block,
    Budget,
    IncomeEntry,
    IncomeSource,
    LineItem,
    PaymentSource,
    Transaction,
    User,
)
from app.security import get_current_user

router = APIRouter(tags=["backup"])

XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ---------------------------------------------------------------- export -----
@router.get("/export")
def export_data(
    db: Session = Depends(get_db), user: User = Depends(get_current_user)
) -> StreamingResponse:
    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet

    # Lookups for id -> name so the sheets are readable.
    blocks = {b.id: b for b in db.scalars(select(Block).where(Block.user_id == user.id))}
    items = {li.id: li for li in db.scalars(select(LineItem).where(LineItem.user_id == user.id))}
    psrc = {p.id: p.name for p in db.scalars(select(PaymentSource).where(PaymentSource.user_id == user.id))}
    isrc = {s.id: s.name for s in db.scalars(select(IncomeSource).where(IncomeSource.user_id == user.id))}

    ws = wb.create_sheet("Blocks")
    ws.append(["Block", "Type"])
    for b in blocks.values():
        ws.append([b.name, b.type])

    ws = wb.create_sheet("Line Items")
    ws.append(["Block", "Line Item"])
    for li in items.values():
        ws.append([blocks[li.block_id].name if li.block_id in blocks else "", li.name])

    ws = wb.create_sheet("Payment Sources")
    ws.append(["Name"])
    for name in psrc.values():
        ws.append([name])

    ws = wb.create_sheet("Income Sources")
    ws.append(["Name"])
    for name in isrc.values():
        ws.append([name])

    ws = wb.create_sheet("Transactions")
    ws.append(["Date", "Amount", "Block", "Line Item", "Payment Source", "Description"])
    for t in db.scalars(
        select(Transaction).where(Transaction.user_id == user.id).order_by(Transaction.txn_date)
    ):
        li = items.get(t.line_item_id)
        block_name = blocks[li.block_id].name if li and li.block_id in blocks else ""
        ws.append([
            t.txn_date.isoformat(),
            float(t.amount),
            block_name,
            li.name if li else "",
            psrc.get(t.payment_source_id, ""),
            t.description or "",
        ])

    ws = wb.create_sheet("Income")
    ws.append(["Date", "Amount", "Source", "Description"])
    for e in db.scalars(
        select(IncomeEntry).where(IncomeEntry.user_id == user.id).order_by(IncomeEntry.entry_date)
    ):
        ws.append([
            e.entry_date.isoformat(),
            float(e.amount),
            isrc.get(e.income_source_id, ""),
            e.description or "",
        ])

    ws = wb.create_sheet("Budgets")
    ws.append(["Block", "Year", "Month", "Amount"])
    for b in db.scalars(select(Budget).where(Budget.user_id == user.id)):
        ws.append([blocks[b.block_id].name if b.block_id in blocks else "", b.year, b.month, float(b.amount)])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"fathom-backup-{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type=XLSX_MEDIA,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------- import -----
def _to_date(value) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value).strip()[:10])


def _rows(wb, sheet: str):
    """Yield data rows (skipping the header) of a sheet if it exists."""
    if sheet not in wb.sheetnames:
        return
    ws = wb[sheet]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or row is None or all(c is None for c in row):
            continue
        yield row


@router.post("/import")
async def import_data(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict:
    try:
        wb = load_workbook(BytesIO(await file.read()), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status.HTTP_400_BAD_REQUEST, f"Not a valid .xlsx file: {exc}")

    # Caches of the current user's entities by name (find-or-create).
    blocks = {b.name: b for b in db.scalars(select(Block).where(Block.user_id == user.id))}
    items: dict[tuple[int, str], LineItem] = {
        (li.block_id, li.name): li
        for li in db.scalars(select(LineItem).where(LineItem.user_id == user.id))
    }
    psrc = {p.name: p for p in db.scalars(select(PaymentSource).where(PaymentSource.user_id == user.id))}
    isrc = {s.name: s for s in db.scalars(select(IncomeSource).where(IncomeSource.user_id == user.id))}

    def get_block(name: str, type_: str = "EXPENSE") -> Block:
        name = str(name).strip()
        if name not in blocks:
            b = Block(user_id=user.id, name=name, type=type_ if type_ in ("EXPENSE", "INVESTMENT") else "EXPENSE")
            db.add(b)
            db.flush()
            blocks[name] = b
        return blocks[name]

    def get_item(block: Block, name: str) -> LineItem:
        name = str(name).strip()
        key = (block.id, name)
        if key not in items:
            li = LineItem(user_id=user.id, block_id=block.id, name=name)
            db.add(li)
            db.flush()
            items[key] = li
        return items[key]

    def get_psrc(name: str) -> PaymentSource:
        name = str(name).strip()
        if name not in psrc:
            p = PaymentSource(user_id=user.id, name=name)
            db.add(p)
            db.flush()
            psrc[name] = p
        return psrc[name]

    def get_isrc(name: str) -> IncomeSource:
        name = str(name).strip()
        if name not in isrc:
            s = IncomeSource(user_id=user.id, name=name)
            db.add(s)
            db.flush()
            isrc[name] = s
        return isrc[name]

    counts = {"transactions": 0, "income": 0, "budgets": 0}
    try:
        # 1) categories/sources first (so later rows can reference them)
        for name, type_ in _rows(wb, "Blocks"):
            if name:
                get_block(name, str(type_ or "EXPENSE"))
        for block_name, item_name in _rows(wb, "Line Items"):
            if block_name and item_name:
                get_item(get_block(block_name), item_name)
        for (name,) in ((r[0],) for r in _rows(wb, "Payment Sources")):
            if name:
                get_psrc(name)
        for (name,) in ((r[0],) for r in _rows(wb, "Income Sources")):
            if name:
                get_isrc(name)

        # 2) transactions
        for row in _rows(wb, "Transactions"):
            d, amount, block_name, item_name, ps_name, desc = (list(row) + [None] * 6)[:6]
            if not (d and amount and block_name and item_name and ps_name):
                continue
            li = get_item(get_block(block_name), item_name)
            db.add(Transaction(
                user_id=user.id, txn_date=_to_date(d), amount=float(amount),
                line_item_id=li.id, payment_source_id=get_psrc(ps_name).id,
                description=(str(desc) if desc else None),
            ))
            counts["transactions"] += 1

        # 3) income
        for row in _rows(wb, "Income"):
            d, amount, src_name, desc = (list(row) + [None] * 4)[:4]
            if not (d and amount and src_name):
                continue
            db.add(IncomeEntry(
                user_id=user.id, entry_date=_to_date(d), amount=float(amount),
                income_source_id=get_isrc(src_name).id,
                description=(str(desc) if desc else None),
            ))
            counts["income"] += 1

        # 4) budgets (upsert)
        for row in _rows(wb, "Budgets"):
            block_name, year, month, amount = (list(row) + [None] * 4)[:4]
            if not (block_name and year and month and amount is not None):
                continue
            block = get_block(block_name)
            existing = db.scalar(select(Budget).where(
                Budget.user_id == user.id, Budget.block_id == block.id,
                Budget.year == int(year), Budget.month == int(month),
            ))
            if existing:
                existing.amount = float(amount)
            else:
                db.add(Budget(user_id=user.id, block_id=block.id, year=int(year), month=int(month), amount=float(amount)))
            counts["budgets"] += 1

        db.commit()
    except HTTPException:
        raise
    except Exception as exc:  # noqa: BLE001
        db.rollback()
        raise HTTPException(status.HTTP_400_BAD_REQUEST, f"Could not import the file: {exc}")

    return {"imported": counts}
