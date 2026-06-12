"""
FileParserService — parse Excel (.xlsx) and CSV bank statements.

Uses openpyxl for Excel and the stdlib csv module for CSV. Both are pure
Python (no C extensions), so installation works the same on Windows,
macOS, and Linux without build tools — no Visual Studio, no compilers.

We deliberately keep this stateless and uncategorized — the route returns
RawStatementRow[] and the frontend's mapping table does category assignment.
That matches the SRS: "No Automated Matching."
"""

import csv
import io
import uuid
from datetime import date, datetime
from typing import Iterable

from openpyxl import load_workbook

from app.schemas import RawStatementRow


_DATE_HINTS = ("date", "txn date", "transaction date", "value date", "posting date")
_AMOUNT_HINTS = ("amount", "debit", "withdrawal", "value", "amt")
_DESC_HINTS = ("description", "narration", "particulars", "details", "remarks", "memo")


def _pick_column(headers: Iterable, hints: tuple[str, ...]) -> int | None:
    """Index of the first header whose name contains any hint substring."""
    for i, col in enumerate(headers):
        if col is None:
            continue
        lower = str(col).strip().lower()
        for hint in hints:
            if hint in lower:
                return i
    return None


def _parse_amount(value) -> float | None:
    """Normalize amounts like '1,234.50', '(500)', '₹ 26,000' → float."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return None
    negative = s.startswith("(") and s.endswith(")")
    cleaned = (
        s.replace("(", "")
        .replace(")", "")
        .replace(",", "")
        .replace("₹", "")
        .replace("$", "")
        .replace("€", "")
        .replace("£", "")
        .strip()
    )
    try:
        amount = float(cleaned)
    except ValueError:
        return None
    return -abs(amount) if negative else amount


def _format_date(value) -> str:
    """Format various date inputs to ISO YYYY-MM-DD, or pass strings through."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _read_csv_rows(contents: bytes) -> list[list]:
    """Decode bytes (handling UTF-8 BOM + Latin-1 fallback) and parse to rows."""
    try:
        text = contents.decode("utf-8-sig")  # utf-8-sig transparently strips BOM
    except UnicodeDecodeError:
        text = contents.decode("latin-1")
    reader = csv.reader(io.StringIO(text))
    return [row for row in reader if any(str(c).strip() for c in row)]


def _read_xlsx_rows(contents: bytes) -> list[list]:
    """Read the active sheet of an .xlsx as a list of row-lists."""
    wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    ws = wb.active
    rows: list[list] = []
    for row in ws.iter_rows(values_only=True):
        # Skip rows that are entirely blank
        if not any(c is not None and str(c).strip() != "" for c in row):
            continue
        rows.append(list(row))
    return rows


def parse_statement(filename: str, contents: bytes) -> list[RawStatementRow]:
    """
    Parse a CSV or XLSX statement into raw, uncategorized rows.

    Returns rows where:
      - date is a string in the source's format (frontend tolerates this)
      - amount is positive (the +/- sign comes from the mapped block type
        on the frontend; bank statements are inconsistent about sign anyway)
      - description is whatever narration column was detected
    """
    name = filename.lower()
    if name.endswith((".xlsx", ".xlsm")):
        rows = _read_xlsx_rows(contents)
    elif name.endswith(".xls"):
        # Legacy .xls is a binary format openpyxl doesn't read. Be explicit.
        raise ValueError(
            "Legacy .xls files aren't supported. Please re-save as .xlsx or export to CSV."
        )
    else:
        rows = _read_csv_rows(contents)

    if len(rows) < 2:
        return []

    headers = rows[0]
    date_idx = _pick_column(headers, _DATE_HINTS)
    amount_idx = _pick_column(headers, _AMOUNT_HINTS)
    desc_idx = _pick_column(headers, _DESC_HINTS)

    # Positional fallbacks when the header names don't match our hints.
    if date_idx is None:
        date_idx = 0
    if amount_idx is None:
        amount_idx = 1 if len(headers) > 1 else None
    if desc_idx is None:
        desc_idx = 2 if len(headers) > 2 else None

    if amount_idx is None:
        return []

    out: list[RawStatementRow] = []
    for row in rows[1:]:
        if amount_idx >= len(row):
            continue
        amount = _parse_amount(row[amount_idx])
        if amount is None:
            # Skip totals rows, header repeats, blank lines — anything where
            # the amount cell isn't a valid number.
            continue

        date_str = _format_date(row[date_idx]) if date_idx < len(row) else ""
        desc = ""
        if desc_idx is not None and desc_idx < len(row):
            v = row[desc_idx]
            if v is not None:
                desc = str(v).strip()

        out.append(
            RawStatementRow(
                row_id=uuid.uuid4().hex,
                date=date_str,
                amount=abs(amount),
                description=desc,
            )
        )
    return out